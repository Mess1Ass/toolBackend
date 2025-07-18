import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font

def save_excel(successful_bids_data, item_name, output_stream, output_file="bidding_results.xlsx"):
    # 保存数据到 Excel 文件
    df = pd.DataFrame(successful_bids_data)
    
    wb = Workbook()
    if not wb.worksheets:
        wb.create_sheet("Sheet1")  # 如果没有工作表就创建一个

    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Sheet1")

    ws.title = "Bidding Results"

    # 写入商品名称
    ws.append([output_file])
    
    # 写入表头并加粗
    header = ["出价状态", "出价人", "出价时间", "出价金额", "座位类型", "座位号"]
    ws.append(header)
    for cell in ws[2]:  # 第二行是标题
        cell.font = Font(bold=True)
    
    # 写入竞价数据
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    
    #更新最小、最大出价、最早、最晚出价的出价人信息
    if(len(successful_bids_data) != 0):
        ws = update_min_max_info(df, ws) 

    wb.save(output_stream)
    print(f"竞价成功信息已保存至 {output_file}" + ".xlsx")
    

def update_min_max_info(df, ws):
    # 只选择出价状态为 "竞价成功" 的记录
    df_successful = df[df['出价状态'] == '竞价成功']

    # 将 '出价时间' 列转换为 datetime 类型，处理格式问题
    # df_successful['出价时间'] = pd.to_datetime(df_successful['出价时间'], format='%Y/%m/%d %H:%M:%S', errors='coerce')
    # df_successful['出价金额'] = pd.to_numeric(df_successful['出价金额'], errors='coerce')
    df_successful.loc[:, '出价时间'] = pd.to_datetime(df_successful['出价时间'], errors='coerce')
    df_successful.loc[:, '出价金额'] = pd.to_numeric(df_successful['出价金额'], errors='coerce')

    # 提取最值
    earliest_bid_time = df_successful['出价时间'].min()
    latest_bid_time = df_successful['出价时间'].max()
    highest_bid = df_successful['出价金额'].max()
    lowest_bid = df_successful['出价金额'].min()
    print(earliest_bid_time, latest_bid_time, highest_bid, lowest_bid)

    # 获取相应的人的详细信息
    earliest_bidder = df_successful[df_successful['出价时间'] == earliest_bid_time].iloc[0]
    latest_bidder = df_successful[df_successful['出价时间'] == latest_bid_time].iloc[0]
    highest_bidder = df_successful[df_successful['出价金额'] == highest_bid].iloc[0]
    lowest_bidder = df_successful[df_successful['出价金额'] == lowest_bid].iloc[0]


    # 将这些值插入到Excel的最后四行
    row_index = len(df) + 3  # 在最后四行之前插入

    # 设置加粗
    bold_font = Font(bold=True)

    ws[f'A{row_index + 1}'].font = bold_font
    ws[f'A{row_index + 2}'].font = bold_font
    ws[f'A{row_index + 3}'].font = bold_font
    ws[f'A{row_index + 4}'].font = bold_font

    # 插入每个人的详细数据，最早出价的那个人的详细数据
    row_number = row_index + 1
    ws[f'A{row_number}'] = '最早出价者'
    ws[f'B{row_number}'] = earliest_bidder['出价人']
    ws[f'C{row_number}'] = earliest_bidder['出价时间'].strftime('%Y-%m-%d %H:%M:%S')
    ws[f'D{row_number}'] = earliest_bidder['出价金额']
    ws[f'E{row_number}'] = earliest_bidder['座位类型']
    ws[f'F{row_number}'] = earliest_bidder['座位号']

    # 插入最晚出价的那个人的详细数据
    row_number += 1
    ws[f'A{row_number}'] = '最晚出价者'
    ws[f'B{row_number}'] = latest_bidder['出价人']
    ws[f'C{row_number}'] = latest_bidder['出价时间'].strftime('%Y-%m-%d %H:%M:%S')
    ws[f'D{row_number}'] = latest_bidder['出价金额']
    ws[f'E{row_number}'] = latest_bidder['座位类型']
    ws[f'F{row_number}'] = latest_bidder['座位号']

    # 插入最高出价的那个人的详细数据
    row_number += 1
    ws[f'A{row_number}'] = '最高出价者'
    ws[f'B{row_number}'] = highest_bidder['出价人']
    ws[f'C{row_number}'] = highest_bidder['出价时间'].strftime('%Y-%m-%d %H:%M:%S')
    ws[f'D{row_number}'] = highest_bidder['出价金额']
    ws[f'E{row_number}'] = highest_bidder['座位类型']
    ws[f'F{row_number}'] = highest_bidder['座位号']

    # 插入最低出价的那个人的详细数据
    row_number += 1
    ws[f'A{row_number}'] = '最低出价者'
    ws[f'B{row_number}'] = lowest_bidder['出价人']
    ws[f'C{row_number}'] = lowest_bidder['出价时间'].strftime('%Y-%m-%d %H:%M:%S')
    ws[f'D{row_number}'] = lowest_bidder['出价金额']
    ws[f'E{row_number}'] = lowest_bidder['座位类型']
    ws[f'F{row_number}'] = lowest_bidder['座位号']

    return ws